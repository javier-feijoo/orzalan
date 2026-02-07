from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook

from db.models import ProductCategory
from db.session import get_session
from paths import get_portable_dir


HEADERS = ["Codigo", "Categoria"]


def export_categories_csv(path: Path) -> Path:
    exports_dir = get_portable_dir("exports")
    if path.is_dir():
        path = exports_dir / "categorias.csv"
    elif not path.is_absolute():
        path = exports_dir / path

    cats = _load_categories()
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        for code, name in cats:
            writer.writerow([code, name])
    return path


def export_categories_xlsx(path: Path) -> Path:
    exports_dir = get_portable_dir("exports")
    if path.is_dir():
        path = exports_dir / "categorias.xlsx"
    elif not path.is_absolute():
        path = exports_dir / path

    cats = _load_categories()
    wb = Workbook()
    ws = wb.active
    ws.title = "Categorias"
    ws.append(HEADERS)
    for code, name in cats:
        ws.append([code, name])
    wb.save(path)
    return path


def import_categories(path: Path) -> tuple[int, int]:
    if not path.exists():
        raise FileNotFoundError(str(path))

    ext = path.suffix.lower()
    if ext == ".csv":
        rows = _read_csv(path)
    elif ext == ".xlsx":
        rows = _read_xlsx(path)
    else:
        raise ValueError("Formato no soportado")

    inserted = 0
    skipped = 0
    with get_session() as session:
        for code, name in rows:
            if not name and not code:
                continue
            if not name:
                name = "Sin categoria"
            if not code:
                code = _code_from_name(name)
            code = code.upper()
            exists = session.query(ProductCategory).filter(ProductCategory.code == code).first()
            if exists is None:
                session.add(ProductCategory(code=code, name=name))
                inserted += 1
            else:
                skipped += 1
        session.commit()
    return inserted, skipped


def _load_categories() -> list[tuple[str, str]]:
    with get_session() as session:
        return [
            (c.code or "", c.name)
            for c in session.query(ProductCategory)
            .order_by(ProductCategory.code.asc().nullslast(), ProductCategory.name.asc())
            .all()
        ]


def _read_csv(path: Path) -> list[tuple[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows: list[tuple[str, str]] = []
        for row in reader:
            code = (row.get("Codigo") or row.get("codigo") or "").strip()
            name = (row.get("Categoria") or row.get("categoria") or "").strip()
            rows.append((code, name))
        return rows


def _read_xlsx(path: Path) -> list[tuple[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None) or []
    header_map = {str(h).strip().lower(): i for i, h in enumerate(headers) if h is not None}
    idx_code = header_map.get("codigo")
    idx_name = header_map.get("categoria")
    if idx_code is None and idx_name is None:
        return []
    rows: list[tuple[str, str]] = []
    for r in rows_iter:
        code = ""
        name = ""
        if idx_code is not None and idx_code < len(r) and r[idx_code] is not None:
            code = str(r[idx_code]).strip()
        if idx_name is not None and idx_name < len(r) and r[idx_name] is not None:
            name = str(r[idx_name]).strip()
        rows.append((code, name))
    return rows


def _code_from_name(name: str) -> str:
    base = "".join([c for c in name.upper() if c.isalnum()])[:3]
    return base or "CAT"
