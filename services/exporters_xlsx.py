from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, numbers
from openpyxl.styles import Font

from db.models import Client, Product, Quote
from db.session import get_session
from sqlalchemy.orm import selectinload
from paths import get_portable_dir, get_base_dir
from settings import Settings
from ui.i18n import t


def export_quote_xlsx(path: Path, quote_id: int, include_costs: bool = False) -> Path:
    exports_dir = get_portable_dir("exports")
    if path.is_dir():
        path = exports_dir / f"presupuesto_{quote_id}.xlsx"
    elif not path.is_absolute():
        path = exports_dir / path

    with get_session() as session:
        quote = (
            session.query(Quote)
            .options(selectinload(Quote.lines))
            .filter(Quote.id == quote_id)
            .first()
        )
        if quote is None:
            raise ValueError(t("quote_not_found"))
        client = session.get(Client, quote.client_id)
        lines = list(quote.lines)

    wb = Workbook()
    ws = wb.active
    ws.title = t("quote")

    bold = Font(bold=True)
    right = Alignment(horizontal="right")

    ws["A1"] = t("quote")
    ws["A1"].font = bold
    ws["A2"] = t("number")
    ws["B2"] = quote.number
    ws["A3"] = t("date")
    ws["B3"] = str(quote.date)
    ws["A4"] = t("client")
    ws["B4"] = client.name if client else ""

    logo_path = Settings.load().get("logo_path", "")
    logo_file = None
    if logo_path:
        candidate = Path(logo_path)
        if candidate.exists():
            logo_file = candidate
    if logo_file is None:
        default_logo = get_base_dir() / "assets" / "logo_orzalan.png"
        if default_logo.exists():
            logo_file = default_logo
    if logo_file is not None:
        try:
            img = XLImage(str(logo_file))
            img.width = 160
            img.height = 80
            ws.add_image(img, "D1")
        except Exception:
            pass

    header = [
        t("ref"),
        t("description"),
        t("unit"),
        t("quantity"),
        t("sale_price"),
        t("subtotal"),
        f"{t('vat')} %",
        t("total"),
    ]
    ws.append([])
    ws.append(header)
    for cell in ws[ws.max_row]:
        cell.font = bold

    for line in lines:
        ws.append(
            [
                line.ref_snapshot,
                line.desc_snapshot,
                line.unit_snapshot,
                float(line.qty or 0),
                float(line.unit_price_snapshot or 0),
                float(line.line_subtotal or 0),
                float(line.vat or 0) / 100.0,
                float(line.line_total or 0),
            ]
        )
        r = ws.max_row
        for col in (4, 5, 6, 8):
            ws.cell(row=r, column=col).number_format = numbers.FORMAT_NUMBER_00
            ws.cell(row=r, column=col).alignment = right
        ws.cell(row=r, column=7).number_format = numbers.FORMAT_PERCENTAGE_00
        ws.cell(row=r, column=7).alignment = right

    ws.append([])
    ws.append([t("subtotal"), float(quote.subtotal or 0)])
    ws.append([f"{t('vat')} %", float(quote.global_vat or 0) / 100.0])
    ws.append([t("vat"), float(quote.vat_total or 0)])
    ws.append([t("total"), float(quote.total or 0)])
    ws.cell(row=ws.max_row - 2, column=2).alignment = right
    ws.cell(row=ws.max_row - 2, column=2).number_format = numbers.FORMAT_NUMBER_00
    ws.cell(row=ws.max_row - 1, column=2).number_format = numbers.FORMAT_PERCENTAGE_00
    ws.cell(row=ws.max_row - 1, column=2).alignment = right
    ws.cell(row=ws.max_row, column=2).number_format = numbers.FORMAT_NUMBER_00
    ws.cell(row=ws.max_row, column=2).alignment = right

    ws2 = wb.create_sheet(t("summary"))
    ws2["A1"] = t("totals")
    ws2["A1"].font = bold
    ws2["A2"] = t("subtotal")
    ws2["B2"] = float(quote.subtotal or 0)
    ws2["A3"] = t("vat")
    ws2["B3"] = float(quote.vat_total or 0)
    ws2["A4"] = t("total")
    ws2["B4"] = float(quote.total or 0)
    ws2["B2"].number_format = numbers.FORMAT_NUMBER_00
    ws2["B3"].number_format = numbers.FORMAT_NUMBER_00
    ws2["B4"].number_format = numbers.FORMAT_NUMBER_00
    ws2["B2"].alignment = right
    ws2["B3"].alignment = right
    ws2["B4"].alignment = right

    _autosize_columns(ws)
    _autosize_columns(ws2)

    if include_costs and Settings.load().get("mostrar_costes", False):
        ws3 = wb.create_sheet(t("cost"))
        ws3.append([t("ref"), t("description"), t("cost"), t("margin")])
        for cell in ws3[1]:
            cell.font = bold
        with get_session() as session:
            for line in lines:
                cost = ""
                margin = ""
                if line.product_id:
                    product = session.get(Product, line.product_id)
                    if product:
                        cost = float(product.cost or 0)
                        margin = float(product.margin or 0) * 100
                ws3.append([line.ref_snapshot, line.desc_snapshot, cost, margin])
        _autosize_columns(ws3)

    wb.save(path)
    return path


def _autosize_columns(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value is None:
                continue
            val = str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
