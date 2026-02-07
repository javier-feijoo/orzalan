from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from PySide6.QtCore import Qt
from PySide6.QtGui import QStandardItem, QStandardItemModel
from PySide6.QtWidgets import (
    QCheckBox,
    QComboBox,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QMessageBox,
    QPushButton,
    QTableView,
    QVBoxLayout,
    QWizard,
    QWizardPage,
    QHeaderView,
)

from db.models import Product, ProductCategory
from db.session import get_session
from paths import get_portable_dir
from ui.i18n import t, tu
from ui.app_events import app_events
from ui.numeric_delegate import NumericAlignDelegate


@dataclass
class ImportStats:
    inserted: int = 0
    updated: int = 0
    rejected: int = 0


def _field_options() -> list[tuple[str, str]]:
    return [
        ("ref", t("ref")),
        ("category_code", t("category_code")),
        ("category", t("category")),
        ("short_desc", t("description")),
        ("long_desc", t("detail")),
        ("unit", t("unit")),
        ("cost", t("price_cost")),
        ("margin", t("benefit")),
        ("sale_price", t("sale_price")),
        ("fixed_price", t("fixed_price")),
    ]


HEADER_ALIASES: dict[str, list[str]] = {
    "ref": ["referencia", "ref", "codigo", "código"],
    "category_code": ["codigo categoria", "código categoria", "codigo de categoria", "código de categoría"],
    "category": ["categoria", "categoría"],
    "short_desc": ["nombre", "descripcion corta", "descripción corta", "producto"],
    "long_desc": ["descripcion", "descripción", "detalle", "descripcion larga", "descripción larga"],
    "unit": ["unidad", "ud", "unid"],
    "cost": ["precio coste", "coste", "costo", "precio costo"],
    "margin": ["beneficio", "margen", "marxe"],
    "sale_price": ["precio venta", "precio de venta", "pvp"],
    "fixed_price": ["precio fijo", "fixed", "fix", "fijo"],
}


class ImportWizard(QWizard):
    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle(tu("import_wizard"))
        self.setWizardStyle(QWizard.ModernStyle)
        self.resize(900, 600)

        self.file_path: Path | None = None
        self.sheet_name: str | None = None
        self.columns: list[str] = []
        self.rows: list[dict[str, Any]] = []
        self.mapping: dict[str, str | None] = {}
        self.stats = ImportStats()
        self.errors: list[str] = []

        self.addPage(FilePage(self))
        self.addPage(SheetPage(self))
        self.addPage(MappingPage(self))
        self.addPage(PreviewPage(self))
        self.addPage(ImportPage(self))
        app_events.language_changed.connect(self._reload_texts)

    def _reload_texts(self, _lang: str) -> None:
        self.setWindowTitle(tu("import_wizard"))
        for page_id in self.pageIds():
            page = self.page(page_id)
            if hasattr(page, "retranslate"):
                page.retranslate()


class FilePage(QWizardPage):
    def __init__(self, wizard: ImportWizard) -> None:
        super().__init__()
        self.wizard_ref = wizard
        self.setTitle(tu("select_file"))

        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        row = QHBoxLayout()
        self.ed_path = QLineEdit()
        self.ed_path.setReadOnly(True)
        btn = QPushButton("...")
        btn.clicked.connect(self._pick_file)
        row.addWidget(self.ed_path, 1)
        row.addWidget(btn)

        layout.addLayout(row)
        layout.addStretch(1)

    def _pick_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, t("select_file"), "", "CSV (*.csv);;Excel (*.xlsx)"
        )
        if not path:
            return
        self.ed_path.setText(path)
        self.wizard_ref.file_path = Path(path)

    def validatePage(self) -> bool:
        path = self.ed_path.text().strip()
        if not path:
            QMessageBox.warning(self, t("import_title"), t("select_file_short"))
            return False
        p = Path(path)
        if not p.exists():
            QMessageBox.warning(self, t("import_title"), t("file_not_found"))
            return False
        self.wizard_ref.file_path = p
        return True

    def retranslate(self) -> None:
        self.setTitle(tu("select_file"))


class SheetPage(QWizardPage):
    def __init__(self, wizard: ImportWizard) -> None:
        super().__init__()
        self.wizard_ref = wizard
        self.setTitle(tu("sheet_title"))

        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        self.cb_sheet = QComboBox()
        self.lbl_select = QLabel(t("select_sheet"))
        layout.addWidget(self.lbl_select)
        layout.addWidget(self.cb_sheet)
        layout.addStretch(1)

    def initializePage(self) -> None:
        self.cb_sheet.clear()
        path = self.wizard_ref.file_path
        if path is None:
            return
        if path.suffix.lower() != ".xlsx":
            self.cb_sheet.addItem(t("not_applicable"))
            self.cb_sheet.setEnabled(False)
            return
        self.cb_sheet.setEnabled(True)
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True)
        for name in wb.sheetnames:
            self.cb_sheet.addItem(name)

    def validatePage(self) -> bool:
        path = self.wizard_ref.file_path
        if path is None:
            return False
        if path.suffix.lower() == ".xlsx":
            self.wizard_ref.sheet_name = self.cb_sheet.currentText()
        else:
            self.wizard_ref.sheet_name = None
        return True

    def retranslate(self) -> None:
        self.setTitle(tu("sheet_title"))
        self.lbl_select.setText(t("select_sheet"))


class MappingPage(QWizardPage):
    def __init__(self, wizard: ImportWizard) -> None:
        super().__init__()
        self.wizard_ref = wizard
        self.setTitle(tu("mapping_title"))

        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        self.group = QGroupBox(t("mapping_group"))
        form = QFormLayout(self.group)
        form.setLabelAlignment(Qt.AlignRight)
        form.setFormAlignment(Qt.AlignTop)
        form.setHorizontalSpacing(12)
        form.setVerticalSpacing(8)
        self.form = form

        self.combos: dict[str, QComboBox] = {}
        self.form_fields: list[tuple[str, QComboBox]] = []
        for field, label in _field_options():
            cb = QComboBox()
            cb.addItem(t("unmapped"), None)
            self.combos[field] = cb
            form.addRow(label, cb)
            self.form_fields.append((field, cb))

        layout.addWidget(self.group)
        layout.addStretch(1)

    def initializePage(self) -> None:
        path = self.wizard_ref.file_path
        if path is None:
            return

        if path.suffix.lower() == ".xlsx":
            columns, rows = _read_xlsx(path, self.wizard_ref.sheet_name)
        else:
            columns, rows = _read_csv(path)

        self.wizard_ref.columns = columns
        self.wizard_ref.rows = rows

        for cb in self.combos.values():
            cb.clear()
            cb.addItem(t("unmapped"), None)
            for col in columns:
                cb.addItem(str(col), str(col))

        self._autoselect_columns(columns)

    def validatePage(self) -> bool:
        self.wizard_ref.mapping = {k: cb.currentData() for k, cb in self.combos.items()}
        if self.wizard_ref.mapping.get("ref") is None:
            QMessageBox.warning(self, t("import_title"), t("map_ref_required"))
            return False
        return True

    def retranslate(self) -> None:
        self.setTitle(tu("mapping_title"))
        self.group.setTitle(t("mapping_group"))
        for field, cb in self.form_fields:
            label = self.form.labelForField(cb)
            if label is not None:
                for key, text in _field_options():
                    if key == field:
                        label.setText(text)
                        break
            current = cb.currentData()
            cb.blockSignals(True)
            cb.clear()
            cb.addItem(t("unmapped"), None)
            for col in self.wizard_ref.columns:
                cb.addItem(str(col), str(col))
            if current is not None:
                idx = cb.findData(current)
                if idx >= 0:
                    cb.setCurrentIndex(idx)
            cb.blockSignals(False)

    def _autoselect_columns(self, columns: list[str]) -> None:
        normalized = {self._norm_header(c): c for c in columns}
        for field, cb in self.form_fields:
            aliases = HEADER_ALIASES.get(field, [])
            selected: str | None = None
            for alias in aliases:
                key = self._norm_header(alias)
                if key in normalized:
                    selected = normalized[key]
                    break
            if selected is None:
                continue
            idx = cb.findData(selected)
            if idx >= 0:
                cb.setCurrentIndex(idx)

    @staticmethod
    def _norm_header(text: str) -> str:
        return "".join(ch for ch in text.strip().lower() if ch.isalnum())


class PreviewPage(QWizardPage):
    def __init__(self, wizard: ImportWizard) -> None:
        super().__init__()
        self.wizard_ref = wizard
        self.setTitle(tu("preview_title"))

        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        self.model = QStandardItemModel(0, len(_field_options()), self)
        self.model.setHorizontalHeaderLabels([label for _, label in _field_options()])

        self.table = QTableView()
        self.table.setModel(self.model)
        self.table.setItemDelegate(NumericAlignDelegate(self.table))
        self.table.setAlternatingRowColors(True)
        header = self.table.horizontalHeader()
        header.setStretchLastSection(True)
        header.setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.verticalHeader().setVisible(False)
        layout.addWidget(self.table, 1)

        self.list_errors = QListWidget()
        self.lbl_errors = QLabel(t("errors"))
        layout.addWidget(self.lbl_errors)
        layout.addWidget(self.list_errors)

    def initializePage(self) -> None:
        self.model.setRowCount(0)
        self.list_errors.clear()
        rows = self.wizard_ref.rows
        if not rows:
            return
        existing_cats = _get_existing_categories()
        preview = rows[:10]
        for idx, row in enumerate(preview):
            items, errors = self._row_to_items(row, existing_cats)
            self.model.appendRow(items)
            for err in errors:
                self.list_errors.addItem(f"{t('row')} {idx + 1}: {err}")

    def _row_to_items(self, row: dict[str, Any], existing_cats: set[str]) -> tuple[list[QStandardItem], list[str]]:
        errors: list[str] = []
        values: list[QStandardItem] = []
        for field, _label in _field_options():
            col = self.wizard_ref.mapping.get(field)
            val = row.get(col, "") if col else ""
            text = "" if val is None else str(val)

            if field in {"ref", "short_desc"} and not text.strip():
                errors.append(f"{t('required_field')}: {field}")
            item = QStandardItem(text)
            if field in {"ref", "short_desc"} and not text.strip():
                item.setBackground(Qt.red)
            if field in {"cost", "margin", "sale_price"}:
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            if field in {"category", "category_code"}:
                if not text.strip():
                    errors.append(t("category_empty"))
                elif text.strip() not in existing_cats:
                    errors.append(f"{t('category_new')}: '{text.strip()}'")
            values.append(item)
        return values, errors

    def retranslate(self) -> None:
        self.setTitle(tu("preview_title"))
        self.model.setHorizontalHeaderLabels([label for _, label in _field_options()])
        self.lbl_errors.setText(t("errors"))


class ImportPage(QWizardPage):
    def __init__(self, wizard: ImportWizard) -> None:
        super().__init__()
        self.wizard_ref = wizard
        self.setTitle(tu("import_title"))

        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        self.cb_duplicates = QComboBox()
        self.cb_duplicates.addItem(t("dup_skip"), "skip")
        self.cb_duplicates.addItem(t("dup_update"), "update")
        self.cb_duplicates.addItem(t("dup_suffix"), "suffix")

        self.chk_log = QCheckBox(t("save_error_log"))

        self.lbl_strategy = QLabel(t("dup_strategy"))
        layout.addWidget(self.lbl_strategy)
        layout.addWidget(self.cb_duplicates)
        layout.addWidget(self.chk_log)
        layout.addStretch(1)

        self.lbl_summary = QLabel("")
        layout.addWidget(self.lbl_summary)

    def initializePage(self) -> None:
        self.lbl_summary.setText("")

    def validatePage(self) -> bool:
        rows = self.wizard_ref.rows
        if not rows:
            return False
        strategy = self.cb_duplicates.currentData() or "skip"
        stats, errors = _import_rows(rows, self.wizard_ref.mapping, strategy)
        self.wizard_ref.stats = stats
        self.wizard_ref.errors = errors

        self.lbl_summary.setText(
            f"{t('imported')}: {stats.inserted} | {t('updated')}: {stats.updated} | {t('rejected_count')}: {stats.rejected}"
        )

        if self.chk_log.isChecked() and errors:
            _write_error_log(errors)

        return True

    def retranslate(self) -> None:
        self.setTitle(tu("import_title"))
        self.chk_log.setText(t("save_error_log"))
        self.lbl_strategy.setText(t("dup_strategy"))
        current = self.cb_duplicates.currentData()
        self.cb_duplicates.blockSignals(True)
        self.cb_duplicates.clear()
        self.cb_duplicates.addItem(t("dup_skip"), "skip")
        self.cb_duplicates.addItem(t("dup_update"), "update")
        self.cb_duplicates.addItem(t("dup_suffix"), "suffix")
        if current is not None:
            idx = self.cb_duplicates.findData(current)
            if idx >= 0:
                self.cb_duplicates.setCurrentIndex(idx)
        self.cb_duplicates.blockSignals(False)


def _import_rows(
    rows: list[dict[str, Any]], mapping: dict[str, str | None], strategy: str
) -> tuple[ImportStats, list[str]]:
    stats = ImportStats()
    errors: list[str] = []

    with get_session() as session:
        for idx, row in enumerate(rows):
            data = _extract_row(row, mapping)
            if data is None:
                stats.rejected += 1
                errors.append(f"{t('row')} {idx + 1}: {t('required_field')}")
                continue

            ref = data["ref"]
            existing = session.query(Product).filter(Product.ref == ref).first()
            if existing is not None:
                if strategy == "skip":
                    stats.rejected += 1
                    errors.append(f"{t('row')} {idx + 1}: {t('duplicate_ref')} {ref}")
                    continue
                if strategy == "suffix":
                    ref = _unique_ref(session, ref)
                if strategy == "update":
                    _apply_product(existing, data)
                    stats.updated += 1
                    continue

            product = Product(ref=ref)
            _apply_product(product, data)
            session.add(product)
            stats.inserted += 1

        session.commit()
    return stats, errors


def _extract_row(row: dict[str, Any], mapping: dict[str, str | None]) -> dict[str, Any] | None:
    def get(field: str) -> Any:
        col = mapping.get(field)
        if not col:
            return None
        val = row.get(col)
        if val is None or val == "":
            return None
        return val

    ref = _to_str(get("ref"))
    desc = _to_str(get("short_desc"))
    if not ref or not desc:
        return None

    data = {
        "ref": ref,
        "category": _to_str(get("category")) or "",
        "category_code": _to_str(get("category_code")).upper(),
        "short_desc": desc,
        "long_desc": _to_str(get("long_desc")),
        "unit": _to_str(get("unit")),
        "cost": _to_float(get("cost"), 0.0),
        "margin": _normalize_margin(get("margin")),
        "sale_price": _to_float(get("sale_price"), 0.0),
        "fixed_price": _to_str(get("fixed_price")).lower() == "fixed",
    }

    if not data["category"] and not data["category_code"]:
        data["category"] = "Sin categoria"

    if not data["fixed_price"]:
        data["sale_price"] = data["cost"] * (1 + data["margin"])
    return data


def _apply_product(product: Product, data: dict[str, Any]) -> None:
    cat = _ensure_category(data.get("category_code"), data.get("category"))
    product.category_id = cat.id if cat else None
    product.short_desc = data["short_desc"] or ""
    product.long_desc = data["long_desc"] or ""
    product.unit = data["unit"] or ""
    product.cost = data["cost"]
    product.margin = data["margin"]
    product.sale_price = data["sale_price"]
    product.fixed_price = data["fixed_price"]
    product.active = True


def _ensure_category(code: str | None, name: str | None) -> ProductCategory | None:
    with get_session() as session:
        if code:
            cat = session.query(ProductCategory).filter(ProductCategory.code == code).first()
            if cat:
                return cat
        if name:
            cat = session.query(ProductCategory).filter(ProductCategory.name == name).first()
            if cat:
                return cat
        if not name:
            name = "Sin categoria"
        if not code:
            code = _code_from_name(name)
        cat = ProductCategory(code=code, name=name)
        session.add(cat)
        session.commit()
        return cat


def _unique_ref(session, ref: str) -> str:
    counter = 1
    new_ref = f"{ref}-{counter}"
    while session.query(Product).filter(Product.ref == new_ref).first() is not None:
        counter += 1
        new_ref = f"{ref}-{counter}"
    return new_ref


def _normalize_margin(value: Any) -> float:
    if value is None:
        return 0.0
    try:
        val = float(str(value).replace(",", "."))
    except Exception:
        return 0.0
    if val > 1:
        return val / 100.0
    return val


def _to_float(value: Any, default: float) -> float:
    if value is None:
        return default
    try:
        return float(str(value).replace(",", "."))
    except Exception:
        return default


def _to_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = [dict(r) for r in reader]
        return list(reader.fieldnames or []), rows


def _read_xlsx(path: Path, sheet_name: str | None) -> tuple[list[str], list[dict[str, Any]]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return [], []
    columns = [str(h).strip() if h is not None else "" for h in headers]
    rows: list[dict[str, Any]] = []
    for r in rows_iter:
        row = {columns[i]: r[i] if i < len(r) else None for i in range(len(columns))}
        rows.append(row)
    return columns, rows


def _write_error_log(errors: list[str]) -> None:
    exports = get_portable_dir("exports")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = exports / f"import_errors_{ts}.txt"
    path.write_text("\n".join(errors), encoding="utf-8")


def _get_existing_categories() -> set[str]:
    with get_session() as session:
        values = set()
        for c in session.query(ProductCategory).all():
            if c.name:
                values.add(c.name)
            if c.code:
                values.add(c.code)
        return values


def _code_from_name(name: str) -> str:
    base = "".join([c for c in name.upper() if c.isalnum()])[:3]
    return base or "CAT"
